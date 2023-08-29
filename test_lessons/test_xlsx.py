from openpyxl import load_workbook
import os

current_path = os.path.join('resources/file_example_XLSX_50.xlsx')


def test_xlsx():
    workbook = load_workbook(current_path)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'